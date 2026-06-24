"""Preflight diagnostics for RASCAL stage execution."""

from __future__ import annotations

import importlib.util
import json
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rascal.config import ProjectConfig
from rascal.diaad_config import metadata_fields_for_plan, write_diaad_config
from rascal.planner import Plan, create_stage_plan

SEVERITIES = {"ok", "warning", "error"}
KNOWN_PROFILE_KEYS = {
    "blind_columns",
    "branch",
    "branches",
    "diaad_executable",
    "dialog",
    "excluded_speakers",
    "lowercase",
    "metadata_fields",
    "monolog",
    "num_bins",
    "num_coders",
    "prefer_correction",
    "profile",
    "random_seed",
    "reliability_fraction",
    "shuffle_samples",
    "strip_clan",
    "thresholds",
}
EXPECTED_INPUT_FILENAMES_BY_COMMAND = {
    "cus files": ("transcript_tables.xlsx",),
    "cus evaluate": ("cu_coding_by_sample_long.xlsx", "cu_coding_by_utterance.xlsx"),
    "cus analyze": ("cu_coding_by_sample_long.xlsx", "cu_coding_by_utterance.xlsx"),
    "cus rates": ("cu_coding_by_sample_long.xlsx",),
    "powers files": ("transcript_tables.xlsx",),
    "powers evaluate": ("powers_reliability_coding.xlsx",),
    "powers analyze": ("powers_coding.xlsx",),
    "templates times": ("cu_coding_by_utterance.xlsx",),
    "vocab analyze": ("cu_coding_by_utterance.xlsx",),
    "vocab rates": ("cu_coding_by_utterance.xlsx",),
    "words analyze": ("word_counting.xlsx",),
    "words evaluate": ("word_counting.xlsx",),
    "words files": ("cu_coding_by_utterance.xlsx",),
    "words rates": ("word_counting.xlsx", "speaking_times.xlsx"),
}


class PreflightError(ValueError):
    """Raised when preflight diagnostics cannot be produced."""


@dataclass(frozen=True)
class CheckResult:
    """One preflight diagnostic result."""

    code: str
    severity: str
    message: str
    path: str | None = None
    details: dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if self.severity not in SEVERITIES:
            raise PreflightError(f"Unknown preflight severity: {self.severity}")

    def as_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable result."""

        payload: dict[str, Any] = {
            "code": self.code,
            "severity": self.severity,
            "message": self.message,
        }
        if self.path is not None:
            payload["path"] = self.path
        if self.details:
            payload["details"] = self.details
        return payload


@dataclass(frozen=True)
class PreflightReport:
    """A complete preflight report for one stage."""

    profile: str
    branch: str
    stage_id: str
    stage_name: str
    results: tuple[CheckResult, ...]

    @property
    def counts(self) -> dict[str, int]:
        """Return counts by severity."""

        return {
            severity: sum(result.severity == severity for result in self.results)
            for severity in ("ok", "warning", "error")
        }

    @property
    def has_errors(self) -> bool:
        """Return whether any diagnostic is an error."""

        return any(result.severity == "error" for result in self.results)

    def as_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable report."""

        return {
            "profile": self.profile,
            "branch": self.branch,
            "stage_id": self.stage_id,
            "stage_name": self.stage_name,
            "counts": self.counts,
            "results": [result.as_dict() for result in self.results],
        }


def _display_path(path: Path, project_root: Path) -> str:
    try:
        return path.resolve().relative_to(project_root.resolve()).as_posix()
    except ValueError:
        return path.resolve().as_posix()


def _result(
    severity: str,
    code: str,
    message: str,
    *,
    path: Path | None = None,
    project_root: Path | None = None,
    details: dict[str, Any] | None = None,
) -> CheckResult:
    display_path = None
    if path is not None:
        display_path = _display_path(path, project_root) if project_root else path.as_posix()
    return CheckResult(
        code=code,
        severity=severity,
        message=message,
        path=display_path,
        details=details or {},
    )


def _branch_config(profile: dict[str, Any], branch: str) -> dict[str, Any]:
    data = profile.get(branch)
    return data if isinstance(data, dict) else {}


def _iter_files(paths: tuple[Path, ...], pattern: str) -> list[Path]:
    found: list[Path] = []
    for path in paths:
        if path.is_file() and path.name == pattern:
            found.append(path)
        elif path.is_dir():
            found.extend(sorted(candidate for candidate in path.rglob(pattern) if candidate.is_file()))
    return found


def _directory_is_empty(path: Path) -> bool:
    return path.is_dir() and not any(path.iterdir())


def _expected_input_filenames(plan: Plan) -> tuple[str, ...]:
    filenames: list[str] = []
    for command in plan.diaad_command_names:
        for filename in EXPECTED_INPUT_FILENAMES_BY_COMMAND.get(command, ()):
            if filename not in filenames:
                filenames.append(filename)
    return tuple(filenames)


def _check_profile_keys(config: ProjectConfig) -> list[CheckResult]:
    unknown = sorted(set(config.resolved_profile) - KNOWN_PROFILE_KEYS)
    if unknown:
        return [
            _result(
                "warning",
                "unknown_profile_keys",
                "Profile contains keys RASCAL does not inspect yet.",
                details={"keys": unknown},
            )
        ]
    return [_result("ok", "profile_keys_known", "Profile keys are recognized.")]


def _check_diaad_dependency(config: ProjectConfig) -> list[CheckResult]:
    executable = str(config.resolved_profile.get("diaad_executable", "diaad"))
    executable_path = shutil.which(executable)
    module_available = importlib.util.find_spec("diaad") is not None
    if executable_path or module_available:
        return [
            _result(
                "ok",
                "diaad_available",
                "DIAAD is available as an executable or importable package.",
                details={
                    "executable": executable,
                    "executable_path": executable_path,
                    "module_available": module_available,
                },
            )
        ]
    return [
        _result(
            "warning",
            "diaad_not_found",
            "DIAAD was not found on PATH or as an importable package.",
            details={"executable": executable},
        )
    ]


def _check_generated_config(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    try:
        generated = write_diaad_config(config, plan)
    except Exception as exc:  # noqa: BLE001 - surfaced as a preflight diagnostic.
        return [
            _result(
                "error",
                "generated_config_unwritable",
                f"Generated DIAAD config could not be written: {exc}",
                path=plan.generated_config_path,
                project_root=config.project_root,
            )
        ]
    return [
        _result(
            "ok",
            "generated_config_writable",
            "Generated DIAAD config was written successfully.",
            path=generated.config_dir,
            project_root=config.project_root,
        )
    ]


def _check_expected_paths(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    results: list[CheckResult] = []
    for path in plan.expected_inputs:
        if not path.exists():
            results.append(
                _result(
                    "error",
                    "expected_input_missing",
                    "Expected input path is missing.",
                    path=path,
                    project_root=config.project_root,
                )
            )
        elif not path.is_dir():
            results.append(
                _result(
                    "error",
                    "expected_input_not_directory",
                    "Expected input path is not a directory.",
                    path=path,
                    project_root=config.project_root,
                )
            )
        elif _directory_is_empty(path) and plan.stage_type != "manual":
            results.append(
                _result(
                    "warning",
                    "expected_input_empty",
                    "Expected input directory is empty.",
                    path=path,
                    project_root=config.project_root,
                )
            )
        else:
            results.append(
                _result(
                    "ok",
                    "expected_input_present",
                    "Expected input path is present.",
                    path=path,
                    project_root=config.project_root,
                )
            )

    for path in plan.expected_outputs:
        if path.exists() and path.is_dir():
            results.append(
                _result(
                    "ok",
                    "expected_output_present",
                    "Expected output path is present.",
                    path=path,
                    project_root=config.project_root,
                )
            )
        elif path.exists():
            results.append(
                _result(
                    "error",
                    "expected_output_not_directory",
                    "Expected output path exists but is not a directory.",
                    path=path,
                    project_root=config.project_root,
                )
            )
        else:
            results.append(
                _result(
                    "warning",
                    "expected_output_missing",
                    "Expected output path does not exist yet.",
                    path=path,
                    project_root=config.project_root,
                )
            )
    return results


def _check_expected_filenames(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    results: list[CheckResult] = []
    for filename in _expected_input_filenames(plan):
        matches = _iter_files(plan.expected_inputs, filename)
        if not matches:
            results.append(
                _result(
                    "warning",
                    "expected_file_missing",
                    f"Expected input file was not found: {filename}",
                    details={"filename": filename},
                )
            )
        elif len(matches) > 1:
            results.append(
                _result(
                    "error",
                    "duplicate_expected_file",
                    f"Multiple expected input files were found: {filename}",
                    details={
                        "filename": filename,
                        "matches": [
                            _display_path(path, config.project_root) for path in matches
                        ],
                    },
                )
            )
        else:
            results.append(
                _result(
                    "ok",
                    "expected_file_present",
                    f"Expected input file is present: {filename}",
                    path=matches[0],
                    project_root=config.project_root,
                    details={"filename": filename},
                )
            )
    return results


def _tokenize_filename(path: Path) -> set[str]:
    return {token for token in re.split(r"[_\-\s]+", path.stem) if token}


def _filename_has_metadata(tokens: set[str], field: str, expected: Any) -> bool:
    if isinstance(expected, list):
        if field == "site":
            return any(token in expected or any(token.startswith(site) for site in expected) for token in tokens)
        return any(token in expected for token in tokens)
    if isinstance(expected, str):
        try:
            return any(re.fullmatch(expected, token) for token in tokens)
        except re.error:
            return False
    return True


def _check_chat_filenames(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    if "transcripts tabularize" not in plan.diaad_command_names:
        return []

    chat_files = _iter_files(plan.expected_inputs, "*.cha")
    if not chat_files:
        return [
            _result(
                "warning",
                "chat_files_missing",
                "No CHAT .cha files were found in the expected input path.",
            )
        ]

    metadata_fields = metadata_fields_for_plan(config, plan)
    results: list[CheckResult] = []
    for chat_file in chat_files:
        tokens = _tokenize_filename(chat_file)
        missing = [
            field
            for field, expected in metadata_fields.items()
            if not _filename_has_metadata(tokens, field, expected)
        ]
        if missing:
            results.append(
                _result(
                    "warning",
                    "chat_filename_unparsed",
                    "CHAT filename does not expose all configured metadata fields.",
                    path=chat_file,
                    project_root=config.project_root,
                    details={"missing_fields": missing},
                )
            )
        else:
            results.append(
                _result(
                    "ok",
                    "chat_filename_parsed",
                    "CHAT filename matches configured metadata fields.",
                    path=chat_file,
                    project_root=config.project_root,
                )
            )
    return results


def _read_sheet_records(sheet) -> tuple[list[str], list[dict[str, Any]]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    records = [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows
    ]
    return headers, records


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _duplicates(values: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicate_values: list[str] = []
    for value in values:
        if value in seen and value not in duplicate_values:
            duplicate_values.append(value)
        seen.add(value)
    return duplicate_values


def _check_required_workbook_structure(
    config: ProjectConfig,
    workbook_path: Path,
    sheets: dict[str, Any],
) -> list[CheckResult]:
    results: list[CheckResult] = []
    required = {
        "samples": {"sample_id"},
        "utterances": {"sample_id", "utterance_id"},
    }
    for sheet_name, required_columns in required.items():
        sheet = sheets.get(sheet_name)
        if sheet is None:
            results.append(
                _result(
                    "error",
                    "transcript_table_sheet_missing",
                    f"Transcript table workbook is missing the {sheet_name!r} sheet.",
                    path=workbook_path,
                    project_root=config.project_root,
                )
            )
            continue
        headers, _ = _read_sheet_records(sheet)
        missing = sorted(required_columns - set(headers))
        if missing:
            results.append(
                _result(
                    "error",
                    "transcript_table_columns_missing",
                    f"Transcript table sheet {sheet_name!r} is missing required columns.",
                    path=workbook_path,
                    project_root=config.project_root,
                    details={"sheet": sheet_name, "missing_columns": missing},
                )
            )
    if not results:
        results.append(
            _result(
                "ok",
                "transcript_table_structure_ok",
                "Transcript table workbook has required sheets and columns.",
                path=workbook_path,
                project_root=config.project_root,
            )
        )
    return results


def _check_sample_ids(
    config: ProjectConfig,
    workbook_path: Path,
    samples: list[dict[str, Any]],
    utterances: list[dict[str, Any]],
) -> list[CheckResult]:
    results: list[CheckResult] = []
    sample_ids = [_clean_value(row.get("sample_id")) for row in samples if _clean_value(row.get("sample_id"))]
    duplicate_samples = _duplicates(sample_ids)
    if duplicate_samples:
        results.append(
            _result(
                "error",
                "duplicate_sample_id",
                "Transcript table samples sheet contains duplicate sample_id values.",
                path=workbook_path,
                project_root=config.project_root,
                details={"sample_id": duplicate_samples},
            )
        )

    utterance_keys = [
        f"{_clean_value(row.get('sample_id'))}:{_clean_value(row.get('utterance_id'))}"
        for row in utterances
        if _clean_value(row.get("sample_id")) and _clean_value(row.get("utterance_id"))
    ]
    duplicate_utterances = _duplicates(utterance_keys)
    if duplicate_utterances:
        results.append(
            _result(
                "error",
                "duplicate_utterance_id",
                "Transcript table utterances sheet contains duplicate sample_id/utterance_id pairs.",
                path=workbook_path,
                project_root=config.project_root,
                details={"keys": duplicate_utterances},
            )
        )

    if not duplicate_samples and not duplicate_utterances:
        results.append(
            _result(
                "ok",
                "sample_identifiers_unique",
                "Sample and utterance identifiers are unique where present.",
                path=workbook_path,
                project_root=config.project_root,
            )
        )
    return results


def _check_value_against_metadata(value: str, expected: Any) -> bool:
    if not value:
        return True
    if isinstance(expected, list):
        return value in {str(item) for item in expected}
    if isinstance(expected, str):
        try:
            return re.fullmatch(expected, value) is not None
        except re.error:
            return True
    return True


def _check_workbook_metadata(
    config: ProjectConfig,
    plan: Plan,
    workbook_path: Path,
    samples_headers: list[str],
    samples: list[dict[str, Any]],
) -> list[CheckResult]:
    results: list[CheckResult] = []
    metadata_fields = metadata_fields_for_plan(config, plan)
    for field, expected in metadata_fields.items():
        if field not in samples_headers:
            results.append(
                _result(
                    "error",
                    "metadata_column_missing",
                    "Transcript table samples sheet is missing a configured metadata column.",
                    path=workbook_path,
                    project_root=config.project_root,
                    details={"column": field},
                )
            )
            continue
        values = sorted(
            {
                _clean_value(row.get(field))
                for row in samples
                if _clean_value(row.get(field))
            }
        )
        unexpected = [
            value
            for value in values
            if not _check_value_against_metadata(value, expected)
        ]
        if unexpected:
            results.append(
                _result(
                    "error",
                    "metadata_value_unexpected",
                    "Transcript table contains metadata values outside the configured profile.",
                    path=workbook_path,
                    project_root=config.project_root,
                    details={"column": field, "values": unexpected},
                )
            )

    if not any(result.severity == "error" and result.code.startswith("metadata_") for result in results):
        results.append(
            _result(
                "ok",
                "metadata_values_ok",
                "Transcript table metadata values match configured fields where present.",
                path=workbook_path,
                project_root=config.project_root,
            )
        )

    if plan.branch == "monolog" and any(command.startswith("vocab ") for command in plan.diaad_command_names):
        corelex = set(_branch_config(config.resolved_profile, "monolog").get("corelex_stimuli", []))
        if corelex and "narrative" in samples_headers:
            narratives = {
                _clean_value(row.get("narrative"))
                for row in samples
                if _clean_value(row.get("narrative"))
            }
            outside_corelex = sorted(narratives - corelex)
            if outside_corelex:
                results.append(
                    _result(
                        "warning",
                        "corelex_subset",
                        "CoreLex-compatible analysis should be restricted to configured CoreLex stimuli.",
                        path=workbook_path,
                        project_root=config.project_root,
                        details={"non_corelex_stimuli": outside_corelex},
                    )
                )
    return results


def _check_transcript_workbooks(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    workbook_paths = sorted(
        set(_iter_files(plan.expected_inputs + plan.expected_outputs, "transcript_tables.xlsx"))
    )
    results: list[CheckResult] = []
    for workbook_path in workbook_paths:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - surfaced as a diagnostic.
            results.append(
                _result(
                    "error",
                    "transcript_table_unreadable",
                    f"Transcript table workbook could not be read: {exc}",
                    path=workbook_path,
                    project_root=config.project_root,
                )
            )
            continue
        try:
            sheets = {name: workbook[name] for name in workbook.sheetnames}
            results.extend(_check_required_workbook_structure(config, workbook_path, sheets))
            if "samples" not in sheets or "utterances" not in sheets:
                continue
            samples_headers, samples = _read_sheet_records(sheets["samples"])
            _, utterances = _read_sheet_records(sheets["utterances"])
            results.extend(_check_sample_ids(config, workbook_path, samples, utterances))
            results.extend(
                _check_workbook_metadata(
                    config,
                    plan,
                    workbook_path,
                    samples_headers,
                    samples,
                )
            )
        finally:
            workbook.close()
    return results


def is_spacy_model_available(model_name: str) -> bool:
    """Return whether spaCy and a named model can be loaded by package name."""

    if importlib.util.find_spec("spacy") is None:
        return False
    try:
        from spacy.util import get_package_path, is_package
    except Exception:  # noqa: BLE001 - absence is reported as unavailable.
        return False
    if is_package(model_name):
        return True
    try:
        get_package_path(model_name)
    except Exception:  # noqa: BLE001 - spaCy raises for missing packages.
        return False
    return True


def _check_spacy(config: ProjectConfig, plan: Plan) -> list[CheckResult]:
    if plan.branch != "dialog":
        return []

    dialog = _branch_config(config.resolved_profile, "dialog")
    if not bool(dialog.get("powers_automation", False)):
        return [
            _result(
                "ok",
                "powers_automation_disabled",
                "POWERS automation is disabled for this profile.",
            )
        ]

    model_name = str(dialog.get("spacy_model", "en_core_web_sm"))
    if is_spacy_model_available(model_name):
        return [
            _result(
                "ok",
                "spacy_model_available",
                "Configured spaCy model is available for POWERS automation.",
                details={"model": model_name},
            )
        ]
    return [
        _result(
            "warning",
            "spacy_model_missing",
            "Configured spaCy model was not found for POWERS automation.",
            details={"model": model_name},
        )
    ]


def _check_thresholds(config: ProjectConfig) -> list[CheckResult]:
    thresholds = config.resolved_profile.get("thresholds")
    if not isinstance(thresholds, dict):
        return [_result("warning", "thresholds_missing", "Reliability thresholds are not configured.")]
    if thresholds.get("report_only") is not True:
        return [
            _result(
                "warning",
                "thresholds_not_report_only",
                "Reliability thresholds are configured as more than report-only.",
            )
        ]
    quality_bands = thresholds.get("quality_bands")
    if not isinstance(quality_bands, dict) or not {"minimal", "sufficient", "excellent"} <= set(quality_bands):
        return [
            _result(
                "warning",
                "threshold_bands_incomplete",
                "Reliability threshold quality bands are incomplete.",
            )
        ]
    return [
        _result(
            "ok",
            "thresholds_report_only",
            "Reliability thresholds are present and report-only.",
            details={"quality_bands": quality_bands},
        )
    ]


def _check_manual_prerequisites(plan: Plan) -> list[CheckResult]:
    if not plan.manual_prerequisites and plan.stage_type != "manual":
        return []
    if plan.stage_type == "manual":
        return [
            _result(
                "warning",
                "manual_stage",
                "This stage is manual; RASCAL cannot confirm completion automatically yet.",
            )
        ]
    return [
        _result(
            "warning",
            "manual_prerequisite_unconfirmed",
            "This automated stage depends on manual work that RASCAL cannot fully verify yet.",
            details={"manual_prerequisites": list(plan.manual_prerequisites)},
        )
    ]


def run_preflight(config: ProjectConfig, branch: str = "common", stage_id: str = "1") -> PreflightReport:
    """Run preflight checks for one branch/stage."""

    plan = create_stage_plan(config, branch, stage_id)
    results: list[CheckResult] = [
        _result("ok", "config_loaded", "RASCAL project config loaded.", path=config.config_path, project_root=config.project_root),
        _result("ok", "profile_loaded", "RASCAL profile resolved.", details={"profile": config.profile_name}),
        _result("ok", "stage_loaded", "RASCAL stage resolved.", details={"stage": f"{branch}:{stage_id}"}),
    ]
    results.extend(_check_profile_keys(config))
    results.extend(_check_diaad_dependency(config))
    results.extend(_check_generated_config(config, plan))
    results.extend(_check_expected_paths(config, plan))
    results.extend(_check_expected_filenames(config, plan))
    results.extend(_check_chat_filenames(config, plan))
    results.extend(_check_transcript_workbooks(config, plan))
    results.extend(_check_spacy(config, plan))
    results.extend(_check_thresholds(config))
    results.extend(_check_manual_prerequisites(plan))
    return PreflightReport(
        profile=config.profile_name,
        branch=plan.branch,
        stage_id=plan.stage_id,
        stage_name=plan.stage_name,
        results=tuple(results),
    )


def render_preflight_json(report: PreflightReport) -> str:
    """Render a preflight report as JSON."""

    return json.dumps(report.as_dict(), indent=2)


def render_preflight_text(report: PreflightReport) -> str:
    """Render a preflight report as concise text."""

    counts = report.counts
    lines = [
        f"RASCAL check: {report.branch} {report.stage_id}",
        f"Profile: {report.profile}",
        f"Stage: {report.stage_name}",
        f"Summary: {counts['error']} error(s), {counts['warning']} warning(s), {counts['ok']} ok",
        "",
        "Results:",
    ]
    for result in report.results:
        path = f" ({result.path})" if result.path else ""
        lines.append(f"  {result.severity.upper()} [{result.code}] {result.message}{path}")
    return "\n".join(lines)
